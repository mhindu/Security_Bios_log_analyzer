import json
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Load the JSON rules
with open('json_final.json', 'r') as file:
    data = json.load(file)

# Access the "checks" key
checks = data.get("checks", [])

# Load keyword mapping from JSON
keyword_mapping = data.get("keyword_mapping", {})

# Function to search for the last occurrence of a keyword in a log file and extract the next bits or value based on JSON rules
def search_last_keyword_in_log(log_file_path, keyword, rule):
    """
    Search for the last occurrence of a keyword in a log file and extract the next value based on JSON rules.
    """
    last_line = None

    with open(log_file_path, 'r') as file:
        lines = file.readlines()

    for line in lines:
        # Normalize line and keyword for case-insensitive comparison
        normalized_line = line.strip().lower()
        if keyword.lower() in normalized_line:
            last_line = line.strip()

    if last_line:
        # Extract the part of the line after the keyword
        start_index = last_line.lower().find(keyword.lower()) + len(keyword) + rule.get("start_index_offset", 0)
        extracted_line = last_line[start_index:].strip()

        # Apply regex pattern from JSON if available
        regex_pattern = rule.get("regex_pattern", "")
        if regex_pattern:
            match = re.search(regex_pattern, extracted_line)
            if match:
                result = match.group(0)  # Extract the matched value
            else:
                result = keyword  # Use the keyword as the result if regex does not match
        else:
            result = extracted_line if extracted_line else keyword  # Use extracted line or keyword as result

        # Remove excluded characters if specified in JSON
        exclude_chars = rule.get("exclude_chars", [])
        for char in exclude_chars:
            result = result.replace(char, '')
        return result

    return None  # Explicitly return None if keyword is not found

def process_last_keyword(log_file_path, keyword, rule):
    """
    Process the last occurrence of a keyword using the search logic.
    """
    result = search_last_keyword_in_log(log_file_path, keyword, rule)
    if result:
        return result.strip()
    return None

# Function to analyze IA32_TME_ACTIVATE MSR
def analyze_ia32_tme_activate(keyword, result, rule, worksheet, row_num):
    """
    Analyze IA32_TME_ACTIVATE MSR dynamically based on JSON rules and update the worksheet.
    """
    if not result:
        worksheet.cell(row=row_num, column=4).value = f"{rule.get('name', 'Keyword')} status is not present in log"
        return

    result = result.strip().replace(" ", "")
    hex_pattern = re.compile(rule.get("pattern", ""))
    
    if not hex_pattern.match(result):
        improper_message = rule.get("improper_print_message", "")
        worksheet.cell(row=row_num, column=4).value = improper_message.format(
            ia32_tme_activate_text=result, ia32_tme_capability_text="N/A"
        )
        return

    try:
        analysis = []
        binary_value = bin(int(result, 16))[2:].zfill(64)
        result_key = "result_text_ia32_tme_activate"

        for output in rule[result_key]:
            bits_range = output.get("bits_range", [])
            if len(bits_range) != 2 or not all(isinstance(b, int) for b in bits_range):
                continue

            start = bits_range[0] if bits_range[0] >= 0 else len(binary_value) + bits_range[0]
            end = bits_range[1] if bits_range[1] >= 0 else len(binary_value) + bits_range[1]
            hex_value = binary_value[start:end]

            if not hex_value or not re.match(r"^[01]+$", hex_value):
                raise ValueError(f"Invalid or empty hex_value extracted: '{hex_value}'")

            bits_as_int = int(hex_value, 2)
            hex_value_formatted = f"0x{bits_as_int:X}"

            if output.get("calculation"):
                calculated_value = eval(output["calculation"].replace("bits", str(bits_as_int)))
                analysis.append(output["output_format"].format(
                    hex_value=hex_value_formatted,
                    calculated_value=calculated_value,
                    bits_as_int=bits_as_int
                ))
            elif "algorithm_map" in output:
                algorithm = output["algorithm_map"].get(hex_value_formatted, "Unknown algorithm")
                analysis.append(output["output_format"].format(
                    hex_value=hex_value_formatted,
                    algorithm=algorithm,
                    bits_as_int=bits_as_int
                ))
            else:
                bits_as_int_minus_1 = bits_as_int - 1 if bits_as_int > 0 else 0
                analysis.append(output["output_format"].format(
                    hex_value=hex_value_formatted,
                    total_keys_text=bits_as_int_minus_1,
                    bits_as_int=bits_as_int,
                    bits_as_int_minus_1=bits_as_int_minus_1
                ))

        worksheet.cell(row=row_num, column=4).value = "\n".join(analysis)
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

    except ValueError as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing result: {result} ({str(e)})"

# Function to analyze IA32_TME_CAPABILITY MSR
def analyze_ia32_tme_capability(keyword, result, rule, worksheet, row_num):
    """
    Analyze IA32_TME_CAPABILITY MSR dynamically based on JSON rules and update the worksheet.
    """
    if not result:
        worksheet.cell(row=row_num, column=4).value = f"{rule.get('name', 'Keyword')} status is not present in log"
        return

    result = result.strip().replace(" ", "")
    hex_pattern = re.compile(rule.get("pattern", ""))
    
    if not hex_pattern.match(result):
        improper_message = rule.get("improper_print_message", "")
        worksheet.cell(row=row_num, column=4).value = improper_message.format(
            ia32_tme_activate_text="N/A", ia32_tme_capability_text=result
        )
        return

    try:
        analysis = []
        binary_value = bin(int(result, 16))[2:].zfill(64)
        result_key = "result_text_ia32_tme_capability"

        for output in rule[result_key]:
            bits_range = output.get("bits_range", [])
            if len(bits_range) != 2 or not all(isinstance(b, int) for b in bits_range):
                continue

            start = bits_range[0] if bits_range[0] >= 0 else len(binary_value) + bits_range[0]
            end = bits_range[1] if bits_range[1] >= 0 else len(binary_value) + bits_range[1]
            hex_value = binary_value[start:end]

            if not hex_value or not re.match(r"^[01]+$", hex_value):
                raise ValueError(f"Invalid or empty hex_value extracted: '{hex_value}'")

            bits_as_int = int(hex_value, 2)
            hex_value_formatted = f"0x{bits_as_int:X}"

            if output.get("calculation"):
                calculated_value = eval(output["calculation"].replace("bits", str(bits_as_int)))
                analysis.append(output["output_format"].format(
                    hex_value=hex_value_formatted,
                    calculated_value=calculated_value,
                    bits_as_int=bits_as_int
                ))
            elif "algorithm_map" in output:
                algorithm = output["algorithm_map"].get(hex_value_formatted, "Unknown algorithm")
                analysis.append(output["output_format"].format(
                    hex_value=hex_value_formatted,
                    algorithm=algorithm,
                    bits_as_int=bits_as_int
                ))
            else:
                bits_as_int_minus_1 = bits_as_int - 1 if bits_as_int > 0 else 0
                analysis.append(output["output_format"].format(
                    hex_value=hex_value_formatted,
                    total_keys_text=bits_as_int_minus_1,
                    bits_as_int=bits_as_int,
                    bits_as_int_minus_1=bits_as_int_minus_1
                ))

        worksheet.cell(row=row_num, column=4).value = "\n".join(analysis)
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

    except ValueError as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing result: {result} ({str(e)})"

# Function to analyze MKTME: total keys
def analyze_total_keys(keyword, result, rule, worksheet, row_num):
    """
    Analyze MKTME: total keys dynamically based on JSON rules and update the worksheet.
    """
    if not result:
        worksheet.cell(row=row_num, column=4).value = f"{rule.get('name', 'Keyword')} status is not present in log"
        return

    try:
        analysis = []
        result_key = "result_text_total_keys"

        # Extract total_keys_text from the result
        total_keys_text = int(result)  # Convert the result to an integer

        for output in rule[result_key]:
            if "output_format" in output:
                analysis.append(output["output_format"].format(
                    total_keys_text=total_keys_text
                ))

        worksheet.cell(row=row_num, column=4).value = "\n".join(analysis)
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

    except ValueError as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing result: {result} ({str(e)})"

# Function to analyze mktme key-ids
def analyze_mktme_keys(keyword, result, rule, worksheet, row_num):
    """
    Analyze mktme key-ids dynamically based on JSON rules and update the worksheet.
    """
    if not result:
        worksheet.cell(row=row_num, column=4).value = f"{rule.get('name', 'Keyword')} status is not present in log"
        return

    result = result.strip().replace(" ", "")
    hex_pattern = re.compile(rule.get("pattern", ""))
    
    if not hex_pattern.match(result):
        worksheet.cell(row=row_num, column=4).value = "Invalid result format"
        return

    try:
        analysis = []
        result_key = "result_text_mktme_keys"
        total_keys_text = worksheet.cell(row=row_num - 1, column=3).value
        mktme_key_text = result

        for output in rule[result_key]:
            if "condition" in output:
                condition = output["condition"]
                condition_evaluated = eval(condition.replace("total_keys_text", str(total_keys_text)).replace("mktme_key_text", str(mktme_key_text)))
                if condition_evaluated:
                    analysis.append(output["output_format"].format(
                        total_keys_text=total_keys_text,
                        mktme_key_text=mktme_key_text
                    ))
            else:
                analysis.append(output["output_format"].format(
                    total_keys_text=total_keys_text,
                    mktme_key_text=mktme_key_text
                ))

        worksheet.cell(row=row_num, column=4).value = "\n".join(analysis)
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

    except ValueError as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing result: {result} ({str(e)})"

# Function to analyze ACTM location
def analyze_actm_location(keyword, result, rule, worksheet, row_num):
    """
    Analyze ACTM location dynamically based on JSON rules and update the worksheet.
    """
    if not result:
        worksheet.cell(row=row_num, column=4).value = rule.get("improper_print_message", "ACTM location is not present in log")
        return

    result = result.strip().replace(" ", "")
    validation_pattern = rule.get("validation_pattern", "")
    
    if not re.match(validation_pattern, result):
        validation_failure_message = rule.get("validation_failure_message", "Invalid ACTM location format")
        worksheet.cell(row=row_num, column=4).value = validation_failure_message.format(actm_location_text=result)
        return

    try:
        # Write the ACTM location to the worksheet
        worksheet.cell(row=row_num, column=3).value = result  # Column C for actm_location_text
        worksheet.cell(row=row_num, column=4).value = f"ACTM location: {result}"  # Column D for analysis
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

    except ValueError as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing ACTM location: {result} ({str(e)})"

# Function to analyze start ACTM
def analyze_start_actm(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze the start ACTM status dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B7", "Is ACTM launch started?")  # Column B

    if result:
        worksheet.cell(row=row_num, column=3).value = rule["output_cells"].get("C7", "ACTM launch started")  # Column C
        worksheet.cell(row=row_num, column=4).value = rule["output_cells"].get("D7", "ACTM launch started")  # Column D
    else:
        secondary_keyword = rule.get("secondary_keyword", "")
        secondary_result = search_last_keyword_in_log(log_file_path, secondary_keyword, rule) if secondary_keyword else None
        if secondary_result:
            worksheet.cell(row=row_num, column=3).value = rule.get("secondary_output", "ACTM launch is skipped")  # Column C
            worksheet.cell(row=row_num, column=4).value = rule.get("secondary_output", "ACTM launch is skipped")  # Column D
        else:
            worksheet.cell(row=row_num, column=3).value = rule["default_result"].get("C7", "ACTM launch status is not present in log")  # Column C
            worksheet.cell(row=row_num, column=4).value = rule["default_result"].get("D7", "ACTM launch status is not present in log")  # Column D

    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

# Function to analyze ACTM error code
def analyze_actm_error_code(keyword, result, rule, worksheet, row_num):
    """
    Analyze the ACTM error code dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = "ACTM ErrorCode"  # Column B

    if not result:
        worksheet.cell(row=row_num, column=3).value = rule["default_result"].get("C8", "0x0")  # Default value for Column C
        worksheet.cell(row=row_num, column=4).value = rule["default_result"].get("D8", "ACTM error code is not present in log")  # Column D
        return

    try:
        # Write the ACTM error code to the worksheet
        worksheet.cell(row=row_num, column=3).value = result  # Column C for actm_error_code_text
        worksheet.cell(row=row_num, column=4).value = f"ACTM ErrorCode: {result}"  # Column D for analysis
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

    except ValueError as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing ACTM error code: {result} ({str(e)})"

# Function to analyze PROD SKU
def analyze_prod_sku(keyword, result, rule, worksheet, row_num):
    """
    Analyze the PROD SKU status dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B9", "Is this PROD SKU?")  # Column B

    if not result:
        worksheet.cell(row=row_num, column=3).value = rule.get("improper_print_message", "Prod SKU status is not present in log")  # Column C
        worksheet.cell(row=row_num, column=4).value = rule.get("improper_print_message", "Prod SKU status is not present in log")  # Column D
        return

    result = result.strip().replace(" ", "")
    validation_pattern = rule.get("validation_pattern", "")
    
    if not re.match(validation_pattern, result):
        worksheet.cell(row=row_num, column=3).value = result  # Column C
        worksheet.cell(row=row_num, column=4).value = rule.get("validation_failure_message", "Invalid PROD SKU format").format(prod_sku_text=result)  # Column D
        return

    try:
        # Convert the result to binary and check the bit specified in the rule
        binary_value = bin(int(result, 16))[2:].zfill(32)
        bit_index = rule["result_text"][0]["bit_check"]
        bit_status = binary_value[-(bit_index + 1)]  # Get the bit value (0 or 1)

        if bit_status == "1":
            bit_status_text = rule["result_text"][0]["bit_set_text"]
            sku_type = rule["result_text"][0]["bit_set_type"]
            patch_type = rule["result_text"][0]["bit_set_patch"]
        else:
            bit_status_text = rule["result_text"][0]["bit_unset_text"]
            sku_type = rule["result_text"][0]["bit_unset_type"]
            patch_type = rule["result_text"][0]["bit_unset_patch"]

        # Format the output
        output = rule["result_text"][0]["output_format"].format(
            bit31_status=bit_status_text,
            sku_type=sku_type,
            patch_type=patch_type
        )
        worksheet.cell(row=row_num, column=3).value = result  # Column C
        worksheet.cell(row=row_num, column=4).value = output  # Column D
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

    except Exception as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing PROD SKU: {result} ({str(e)})"

# Function to analyze Mcheck Error Code
def mcheck_error_code(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze Mcheck Error Code dynamically based on JSON rules and update the worksheet.
    """
    # Set header for Mcheck Error Code (Column B)
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B10", "Mcheck Error Code ?")

    # Process primary keyword
    if not result:
        secondary_keyword = rule.get("secondary_keyword", "")
        if secondary_keyword:
            result = process_last_keyword(log_file_path, secondary_keyword, rule)
            if not result:
                worksheet.cell(row=row_num, column=3).value = "0x0"
                worksheet.cell(row=row_num, column=4).value = rule.get("secondary_output", "No Mcheck - SGX enabled successfully")
                return
        else:
            worksheet.cell(row=row_num, column=3).value = "0x0"
            worksheet.cell(row=row_num, column=4).value = rule.get("secondary_output", "No Mcheck - SGX enabled successfully")
            return

    # Apply trim processing if specified
    trim_processing = rule.get("trim_processing", {})
    if trim_processing.get("trim_whitespace", False):
        result = result.strip()
    if trim_processing.get("if_starts_with") and result.startswith(trim_processing["if_starts_with"]):
        replacement_prefix = trim_processing.get("replacement_prefix", "")
        remove_chars_count = trim_processing.get("remove_chars_count", 0)
        result = replacement_prefix + result[remove_chars_count:]

    # Validate the processed result
    validation_pattern = rule.get("validation_pattern", "")
    if not re.match(validation_pattern, result):
        worksheet.cell(row=row_num, column=3).value = result
        worksheet.cell(row=row_num, column=4).value = rule.get("validation_failure_message", "improper print - {mcheck_code_text}").format(mcheck_code_text=result)
        return

    # Search for error description in Mcheck_error_codes.xlsx
    error_description = "Unknown error"
    try:
        error_code_file_path = rule["result_text"][0]["error_code_file_path"]
        error_code_column = rule["result_text"][0]["error_code_column"]
        error_description_column = rule["result_text"][0]["error_description_column"]
        error_codes_df = pd.read_excel(error_code_file_path)
        matching_row = error_codes_df[error_codes_df.iloc[:, error_code_column] == result]
        if not matching_row.empty:
            error_description = matching_row.iloc[0, error_description_column]
    except Exception:
        pass

    # Update worksheet with result and analysis
    worksheet.cell(row=row_num, column=3).value = result
    worksheet.cell(row=row_num, column=4).value = f"Mcheck Error is {error_description}"

# Function to analyze SGX enabled status
def sgx_enabled_check(keyword, result, rule, worksheet, row_num):
    """
    Analyze SGX enabled status dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B11", "Is SGX enabled?")  # Column B

    if result:
        worksheet.cell(row=row_num, column=3).value = result  # Column C
        worksheet.cell(row=row_num, column=4).value = "Pass: SGX feature is enabled"  # Column D
    else:
        worksheet.cell(row=row_num, column=3).value = "Not Found"  # Column C
        worksheet.cell(row=row_num, column=4).value = "Fail: SGX feature is not enabled"  # Column D

# Function to process TDX build date check
def tdx_build_date_check(keyword, result, rule, worksheet, row_num):
    """
    Process the TDX build date check dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B12", "TDX Build Date and version")  # Column B

    if not result:
        worksheet.cell(row=row_num, column=3).value = rule.get("improper_print_message", "TDX build date is not present in log")  # Column C
        worksheet.cell(row=row_num, column=4).value = rule.get("improper_print_message", "TDX build date is not present in log")  # Column D
    else:
        worksheet.cell(row=row_num, column=3).value = result  # Column C
        worksheet.cell(row=row_num, column=4).value = result  # Column D

def tdx_module_initialization_check(keyword, rule, worksheet, row_num, log_file_path):
    """
    Process the TDX module initialization check dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B13", "Is TDX module initialized?")  # Column B

    # Search for the keyword in the log file
    result = search_last_keyword_in_log(log_file_path, keyword, rule)
    if result:
        worksheet.cell(row=row_num, column=3).value = "TDX module is initialized"  # Column C
        worksheet.cell(row=row_num, column=4).value = "TDX module is initialized"  # Column D
    else:
        worksheet.cell(row=row_num, column=3).value = "TDX module is not initialized"  # Column C
        worksheet.cell(row=row_num, column=4).value = "TDX module is not initialized"  # Column D

    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

def seamrr_initialization_check(keyword, rule, worksheet, row_num, log_file_path):
    """
    Process the SEAMRR initialization check dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B14", "Is SEAMRR Initialized?")  # Column B

    # Search for the keyword in the log file
    result = search_last_keyword_in_log(log_file_path, keyword, rule)
    if result:
        worksheet.cell(row=row_num, column=3).value = result  # Column C
        worksheet.cell(row=row_num, column=4).value = "SEAMRR target are initialized"  # Column D
    else:
        worksheet.cell(row=row_num, column=3).value = "Seamrr target not initialized"  # Column C
        worksheet.cell(row=row_num, column=4).value = "SEAMRR target are not initialized"  # Column D

    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

def seamrr_base_range_check(keyword, rule, worksheet, row_num, log_file_path):
    """
    Process the SEAMRR base range check dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B16", "SEAMRR_BASE MSR 1400h")  # Column B

    # Search for the keyword in the log file
    result = search_last_keyword_in_log(log_file_path, keyword, rule)
    if result:
        cleaned_result = result.replace("]", "").replace("=", "").strip().replace(" ", "")
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')

        if hex_pattern.match(cleaned_result):
            worksheet.cell(row=row_num, column=3).value = cleaned_result  # Column C
            seamrr_base_value = int(cleaned_result, 16)
            if seamrr_base_value & (1 << 3):  # Check if bit 3 is set
                worksheet.cell(row=row_num, column=4).value = "Bit 3 is set, Bios configured SEAMRR_Base"  # Column D
            else:
                worksheet.cell(row=row_num, column=4).value = "Bit 3 is not set, Bios does not configure SEAMRR_Base"  # Column D
        else:
            worksheet.cell(row=row_num, column=3).value = f"improper print - {cleaned_result}"  # Column C
    else:
        worksheet.cell(row=row_num, column=3).value = "SEAMRR Base range is not present in log"  # Column C

    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

def seamrr_mask_range_check(keyword, rule, worksheet, row_num, log_file_path):
    """
    Process the SEAMRR mask range check dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B17", "SEAMRR_MASK MSR 1401h")  # Column B

    # Search for the keyword in the log file
    result = search_last_keyword_in_log(log_file_path, keyword, rule)
    if result:
        cleaned_result = result.replace("]", "").replace("=", "").strip().replace(" ", "")
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')

        if hex_pattern.match(cleaned_result):
            worksheet.cell(row=row_num, column=3).value = cleaned_result  # Column C
            seamrr_mask_value = int(cleaned_result, 16)
            bit10_set = seamrr_mask_value & (1 << 10)
            bit11_set = seamrr_mask_value & (1 << 11)

            if bit10_set and bit11_set:
                worksheet.cell(row=row_num, column=4).value = "bit 10 lock bit and bit 11 valid bit are set"  # Column D
            elif bit10_set:
                worksheet.cell(row=row_num, column=4).value = "only bit 10 lock bit is set"  # Column D
            elif bit11_set:
                worksheet.cell(row=row_num, column=4).value = "only bit 11 valid bit is set"  # Column D
            else:
                worksheet.cell(row=row_num, column=4).value = "neither bit 10 nor bit 11 is set on SEAMRR_MASK"  # Column D
        else:
            worksheet.cell(row=row_num, column=3).value = f"improper print - {cleaned_result}"  # Column C
    else:
        worksheet.cell(row=row_num, column=3).value = "SEAMRR Mask range is not present in log"  # Column C

    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

def process_lt_status_info(keyword, result, rule, worksheet, row_num):
    """
    Process the LT_STATUS info dynamically based on JSON rules and update the worksheet.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B18", "LT_STATUS[0xFED30000]")  # Column B

    if not result:
        worksheet.cell(row=row_num, column=3).value = rule.get("improper_print_message", "LT_status is not present in log")  # Column C
        worksheet.cell(row=row_num, column=4).value = rule.get("improper_print_message", "LT_status is not present in log")  # Column D
        return

    result = result.strip()
    validation_pattern = rule.get("validation_pattern", "")
    
    if not re.match(validation_pattern, result):
        worksheet.cell(row=row_num, column=3).value = result  # Column C
        # Supply both keys in case JSON message uses either placeholder
        worksheet.cell(row=row_num, column=4).value = rule.get("validation_failure_message", "improper print - {lt_status_info_text}").format(lt_status_info_text=result, lt_status_text=result)  # Column D
        return

    try:
        lt_status_value = int(result, 16)
        analysis = []
        for output in rule["result_text"]:
            bit_check = output["bit_check"]
            if lt_status_value & (1 << bit_check):
                analysis.append(output["output_format"].format(
                    bit_position=bit_check,
                    bit_status=output["bit_set_text"],
                    bit_message=output["bit_set_message"]
                ))
            else:
                analysis.append(output["output_format"].format(
                    bit_position=bit_check,
                    bit_status=output["bit_unset_text"],
                    bit_message=output["bit_unset_message"]
                ))
        worksheet.cell(row=row_num, column=3).value = result  # Column C
        worksheet.cell(row=row_num, column=4).value = "\n".join(analysis)  # Column D
        worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)
    except Exception as e:
        worksheet.cell(row=row_num, column=4).value = f"Error processing LT_STATUS info: {result} ({str(e)})"

def analyze_lt_extended_status_info(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze LT_EXTENDED_STATUS info dynamically based on JSON rules and update the worksheet.
    The if/else logic is driven by the JSON rule. JSON placeholders include:
      - {bit0_status}: value from bit_set_text or bit_unset_text
      - {poison_status}: value from bit_set_poison_status or bit_unset_poison_status
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B19", "LT_EXTENDED_STATUS[0xFED30008]")
    
    # Use the generic search function to obtain the status text.
    lt_extended_status_info_text = search_last_keyword_in_log(log_file_path, keyword, rule)
    if lt_extended_status_info_text is None:
        worksheet.cell(row=row_num, column=3).value = rule.get("improper_print_message", "LT_extended_status is not present in log")
        return

    lt_extended_status_info_text = lt_extended_status_info_text.strip()
    hex_pattern = re.compile(rule.get("validation_pattern", "^(0x)?[0-9A-Fa-f]{16}$"))
    if not hex_pattern.match(lt_extended_status_info_text):
        worksheet.cell(row=row_num, column=3).value = f"improper print - {lt_extended_status_info_text}"
        return

    worksheet.cell(row=row_num, column=3).value = lt_extended_status_info_text
    lt_extended_status_info_value = int(lt_extended_status_info_text, 16)
    analysis = []
    # Let the JSON rules drive the conditional text (if/else loops remain in JSON)
    for output in rule["result_text"]:
        bit_check = output.get("bit_check")
        # The replacement dictionary is built based on whether the bit is set:
        rep = {
            "bit0_status": output.get("bit_set_text") if lt_extended_status_info_value & (1 << bit_check) else output.get("bit_unset_text"),
            "poison_status": output.get("bit_set_poison_status") if lt_extended_status_info_value & (1 << bit_check) else output.get("bit_unset_poison_status")
        }
        analysis.append(output["output_format"].format(**rep))
    worksheet.cell(row=row_num, column=4).value = "\n".join(analysis)

def analyze_lt_boot_status_info(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze LT_BOOT_STATUS info dynamically based on JSON rules and update the worksheet.
    The JSON rule should provide:
      - output_cells with key "B20" for header.
      - validation_pattern for hex string.
      - result_text (list with one dict) providing:
            bit30_check, bit31_check, bit59_check, bit62_check, bit63_check,
            bit_set_text, bit_unset_text,
            bit30_set_txt_status, bit30_unset_txt_status,
            bit31_set_btg_status, bit31_unset_btg_status,
            bit59_set_bios_status, bit59_unset_bios_status,
            bit62_set_cpu_error_status, bit62_unset_cpu_error_status,
            bit63_set_sacm_status, bit63_unset_sacm_status,
            and an output_format that uses placeholders:
                {bit30_status}, {txt_status},
                {bit31_status}, {btg_status},
                {bit59_status}, {bios_status},
                {bit62_status}, {cpu_error_status},
                {bit63_status}, {sacm_status}.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B20", "LT_BOOT_STATUS[0xFED300A0]")
    
    lt_boot_status_info_text = search_last_keyword_in_log(log_file_path, keyword, rule)
    if lt_boot_status_info_text is None:
        worksheet.cell(row=row_num, column=3).value = rule.get("improper_print_message", "LT_boot_status is not present in log")
        return

    lt_boot_status_info_text = lt_boot_status_info_text.strip()
    hex_pattern = re.compile(rule.get("validation_pattern", "^(0x)?[0-9A-Fa-f]{16}$"))
    if not hex_pattern.match(lt_boot_status_info_text):
        worksheet.cell(row=row_num, column=3).value = f"improper print - {lt_boot_status_info_text}"
        return

    worksheet.cell(row=row_num, column=3).value = lt_boot_status_info_text
    value = int(lt_boot_status_info_text, 16)
    rep = {}
    # Process bit30
    if value & (1 << rule["result_text"][0].get("bit30_check", 30)):
        rep["bit30_status"] = rule["result_text"][0].get("bit_set_text", "Bit {bit} set").replace("{bit}", "30")
        rep["txt_status"] = rule["result_text"][0].get("bit30_set_txt_status", "success")
    else:
        rep["bit30_status"] = rule["result_text"][0].get("bit_unset_text", "Bit {bit} is unset").replace("{bit}", "30")
        rep["txt_status"] = rule["result_text"][0].get("bit30_unset_txt_status", "not success")
    # Process bit31
    if value & (1 << rule["result_text"][0].get("bit31_check", 31)):
        rep["bit31_status"] = rule["result_text"][0].get("bit_set_text", "Bit {bit} set").replace("{bit}", "31")
        rep["btg_status"] = rule["result_text"][0].get("bit31_set_btg_status", "success")
    else:
        rep["bit31_status"] = rule["result_text"][0].get("bit_unset_text", "Bit {bit} is unset").replace("{bit}", "31")
        rep["btg_status"] = rule["result_text"][0].get("bit31_unset_btg_status", "not success")
    # Process bit59
    if value & (1 << rule["result_text"][0].get("bit59_check", 59)):
        rep["bit59_status"] = rule["result_text"][0].get("bit_set_text", "Bit {bit} set").replace("{bit}", "59")
        rep["bios_status"] = rule["result_text"][0].get("bit59_set_bios_status", "trusted")
    else:
        rep["bit59_status"] = rule["result_text"][0].get("bit_unset_text", "Bit {bit} is unset").replace("{bit}", "59")
        rep["bios_status"] = rule["result_text"][0].get("bit59_unset_bios_status", "untrusted")
    # Process bit62
    if value & (1 << rule["result_text"][0].get("bit62_check", 62)):
        rep["bit62_status"] = rule["result_text"][0].get("bit_set_text", "Bit {bit} set").replace("{bit}", "62")
        rep["cpu_error_status"] = rule["result_text"][0].get("bit62_set_cpu_error_status", "SACM authentication failed")
    else:
        rep["bit62_status"] = rule["result_text"][0].get("bit_unset_text", "Bit {bit} is unset").replace("{bit}", "62")
        rep["cpu_error_status"] = rule["result_text"][0].get("bit62_unset_cpu_error_status", "No CPU error - No SACM authentication failure")
    # Process bit63
    if value & (1 << rule["result_text"][0].get("bit63_check", 63)):
        rep["bit63_status"] = rule["result_text"][0].get("bit_set_text", "Bit {bit} set").replace("{bit}", "63")
        rep["sacm_status"] = rule["result_text"][0].get("bit63_set_sacm_status", "success")
    else:
        rep["bit63_status"] = rule["result_text"][0].get("bit_unset_text", "Bit {bit} is unset").replace("{bit}", "63")
        rep["sacm_status"] = rule["result_text"][0].get("bit63_unset_sacm_status", "execution failed")
    
    output_str = rule["result_text"][0]["output_format"].format(**rep)
    worksheet.cell(row=row_num, column=4).value = output_str
    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

def analyze_lt_error_code_info(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze LT_ERROR_CODE info dynamically based on JSON rules and update the worksheet.
    The JSON rule should provide in 'result_text' a dict with:
      - bit31_check, bit15_check, bit30_check,
      - bit31_set_register_status, bit31_unset_register_status,
      - bit15_set_sacm_status, bit15_unset_sacm_status,
      - bit30_set_error_status, bit30_unset_error_status,
      - and an output_format using placeholders: {bit31_status}, {register_status},
        {bit15_status}, {sacm_status}, {bit30_status}, {error_status}.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B21", "LT_ERROR_CODE[0xFED30328]")
    lt_error_code_info_text = search_last_keyword_in_log(log_file_path, keyword, rule)
    if lt_error_code_info_text is None:
        worksheet.cell(row=row_num, column=3).value = rule.get("improper_print_message", "LT_error_status is not present in log")
        return
    lt_error_code_info_text = lt_error_code_info_text.strip()
    hex_pattern = re.compile(rule.get("validation_pattern", "^(0x)?[0-9A-Fa-f]{16}$"))
    if not hex_pattern.match(lt_error_code_info_text):
        worksheet.cell(row=row_num, column=3).value = f"improper print - {lt_error_code_info_text}"
        return
    worksheet.cell(row=row_num, column=3).value = lt_error_code_info_text
    value = int(lt_error_code_info_text, 16)
    rep = {}
    # Process bit 31 check: if set, further analyze bits 15 and 30
    if value & (1 << rule["result_text"][0].get("bit31_check", 31)):
        rep["bit31_status"] = "set"
        rep["register_status"] = rule["result_text"][0].get("bit31_set_register_status", "valid")
        if value & (1 << rule["result_text"][0].get("bit15_check", 15)):
            rep["bit15_status"] = "set"
            rep["sacm_status"] = rule["result_text"][0].get("bit15_set_sacm_status", "started successfully")
        else:
            rep["bit15_status"] = "unset"
            rep["sacm_status"] = rule["result_text"][0].get("bit15_unset_sacm_status", "didnt start")
        if value & (1 << rule["result_text"][0].get("bit30_check", 30)):
            rep["bit30_status"] = "set"
            rep["error_status"] = rule["result_text"][0].get("bit30_set_error_status", "generated errorcodes")
        else:
            rep["bit30_status"] = "unset"
            rep["error_status"] = rule["result_text"][0].get("bit30_unset_error_status", "not generated errorcodes")
    else:
        rep["bit31_status"] = "unset"
        rep["register_status"] = rule["result_text"][0].get("bit31_unset_register_status", "invalid from bits 30:0")
    output_str = rule["result_text"][0]["output_format"].format(**rep)
    worksheet.cell(row=row_num, column=4).value = output_str
    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

def analyze_lt_crash_info(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze LT_CRASH info dynamically based on JSON rules and update the worksheet.
    The JSON rule should provide:
      - output_cells with key "B22" for header.
      - validation_pattern for a 16-digit hex string.
      - result_text (list with one dict) with keys:
            bit31_check, bit30_check,
            bit_set_text, bit_unset_text,
            bit31_set_crash_status, bit31_unset_crash_status,
            bit30_set_crash_source, bit30_unset_crash_source,
            and an output_format.
    When bit 31 is not set, the function will not check bit 30 and will
    output only the bit 31 information.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B22", "LT_CRASH[0xFED30030]")
    lt_crash_info_text = search_last_keyword_in_log(log_file_path, keyword, rule)
    if lt_crash_info_text is None:
        worksheet.cell(row=row_num, column=3).value = rule.get("improper_print_message", "LT_crash is not present in log")
        return
    lt_crash_info_text = lt_crash_info_text.strip()
    hex_pattern = re.compile(rule.get("validation_pattern", "^(0x)?[0-9A-Fa-f]{16}$"))
    if not hex_pattern.match(lt_crash_info_text):
        worksheet.cell(row=row_num, column=3).value = f"improper print - {lt_crash_info_text}"
        return
    worksheet.cell(row=row_num, column=3).value = lt_crash_info_text
    value = int(lt_crash_info_text, 16)
    rep = {}
    if value & (1 << rule["result_text"][0].get("bit31_check", 31)):
        rep["bit31_status"] = rule["result_text"][0].get("bit_set_text", "set")
        rep["crash_status"] = rule["result_text"][0].get("bit31_set_crash_status", "set")
        if value & (1 << rule["result_text"][0].get("bit30_check", 30)):
            rep["crash_source"] = rule["result_text"][0].get("bit30_set_crash_source", "internal CPU")
        else:
            rep["crash_source"] = rule["result_text"][0].get("bit30_unset_crash_source", "external ACM, MLE")
        output_str = rule["result_text"][0]["output_format"].format(**rep)
    else:
        rep["bit31_status"] = rule["result_text"][0].get("bit_unset_text", "unset")
        rep["crash_status"] = rule["result_text"][0].get("bit31_unset_crash_status", "not seen")
        # When bit 31 is not set, we skip checking bit 30.
        output_str = "Bit 31 is {bit31_status} - LT_Crash is {crash_status}".format(**rep)
    worksheet.cell(row=row_num, column=4).value = output_str
    worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)

def analyze_sacm_info(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze SACM_INFO info dynamically based on JSON rules and update the worksheet.
    This function processes the log for "MSR_BOOT_GUARD_SACM_INFO[0x0000013A]" using the rule.
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B23", "SACM_INFO[0x0000013A]")
    sacm_info_text = search_last_keyword_in_log(log_file_path, "MSR_BOOT_GUARD_SACM_INFO[0x0000013A]", rule)
    if sacm_info_text is None:
        worksheet.cell(row=row_num, column=3).value = "SACM_INFO is not present in log"
    else:
        sacm_info_text = sacm_info_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(sacm_info_text):
            worksheet.cell(row=row_num, column=3).value = sacm_info_text
            sacm_info_value = int(sacm_info_text, 16)
            bit3_set  = bool(sacm_info_value & (1 << 3))
            bit34_set = bool(sacm_info_value & (1 << 34))
            bit32_set = bool(sacm_info_value & (1 << 32))
            bit0_set  = bool(sacm_info_value & (1 << 0))
            bit2_set  = bool(sacm_info_value & (1 << 2))
            bit4_set  = bool(sacm_info_value & (1 << 4))
            bit5_set  = bool(sacm_info_value & (1 << 5))
            bit6_set  = bool(sacm_info_value & (1 << 6))
            result_text = []
            if bit34_set:
                result_text.append("Bit 34 set - LT_SX_EN Fuse is enabled")
            else:
                result_text.append("Bit 34 is unset - LT_SX_EN Fuse is disabled")
            if bit32_set:
                result_text.append("Bit 32 set - BTG is enabled")
                if bit0_set and bit2_set and bit3_set and bit4_set and bit5_set and bit6_set:
                    result_text.append("BTG-5 profile is enabled & TPM Success")
                elif bit0_set and bit2_set and bit3_set and bit5_set and bit6_set:
                    result_text.append("BTG-3 profile is enabled & TPM Success")
                elif bit0_set and bit4_set and bit6_set:
                    result_text.append("BTG-4 profile is enabled")
                else:
                    result_text.append("No BTG profile enabled")
            else:
                result_text.append("Bit 32 is unset - BTG is not enabled")
            result_text_str = "\n".join(result_text)
            worksheet.cell(row=row_num, column=4).value = result_text_str
            worksheet.cell(row=row_num, column=4).alignment = Alignment(wrapText=True)
        else:
            worksheet.cell(row=row_num, column=3).value = f"improper print - {sacm_info_text}"

def analyze_bios_id_info(keyword, result, rule, worksheet, row_num, log_file_path):
    """
    Analyze BIOS ID info dynamically based on JSON rules and update the worksheet.
    This function extracts information using the keyword "Bios ID: ".
    """
    worksheet.cell(row=row_num, column=2).value = rule["output_cells"].get("B24", "BIOS ID")
    bios_info_text = search_last_keyword_in_log(log_file_path, "Bios ID: ", rule)
    if bios_info_text is None:
        worksheet.cell(row=row_num, column=3).value = "Bios_ID is not present in log"
    else:
        worksheet.cell(row=row_num, column=3).value = bios_info_text.strip()

# Analyze the log file using the rules
def analyze_log_file(log_file_path, checks):
    results = []
    for check in checks:
        keyword = check.get('keyword', '')
        result = process_last_keyword(log_file_path, keyword, check)
        results.append((keyword, result, check))
    return results

# Save the results to an Excel file
def save_to_excel(results, output_file, log_file_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Analysis Results"

    # Add headers
    headers = ["Security Features", "Feature check", "Results from log", "Analysis from Result"]
    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Add results
    for row_num, (keyword, result, rule) in enumerate(results, start=2):
        security_feature, feature_check = keyword_mapping.get(keyword, ["Unknown", "Unknown"])
        worksheet.cell(row=row_num, column=1).value = security_feature
        worksheet.cell(row=row_num, column=2).value = feature_check
        worksheet.cell(row=row_num, column=3).value = result

        if rule.get("name") == "IA32_TME_ACTIVATE MSR":
            analyze_ia32_tme_activate(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "IA32_TME_CAPABILITY MSR":
            analyze_ia32_tme_capability(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "MKTME: total keys":
            analyze_total_keys(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "mktme key-ids":
            analyze_mktme_keys(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "ACTM location":
            analyze_actm_location(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "start ACTM":
            analyze_start_actm(keyword, result, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "ACTM error code":
            analyze_actm_error_code(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "prod_sku":
            analyze_prod_sku(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "Mcheck Error Code":
            mcheck_error_code(keyword, result, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "SGX enabled check":
            sgx_enabled_check(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "TDX build date check":
            tdx_build_date_check(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "TDX module initialization check":
            tdx_module_initialization_check(keyword, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "SEAMRR initialization check":
            seamrr_initialization_check(keyword, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "SEAMRR base range check":
            seamrr_base_range_check(keyword, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "SEAMRR mask range check":
            seamrr_mask_range_check(keyword, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "LT_STATUS info":
            process_lt_status_info(keyword, result, rule, worksheet, row_num)
        elif rule.get("name") == "LT_EXTENDED_STATUS info":
            analyze_lt_extended_status_info(keyword, result, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "LT_BOOT_STATUS info":
            analyze_lt_boot_status_info(keyword, result, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "LT_ERROR_CODE info":
            analyze_lt_error_code_info(keyword, result, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "LT_CRASH info":
            analyze_lt_crash_info(keyword, result, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "SACM_INFO info":
            analyze_sacm_info(keyword, result, rule, worksheet, row_num, log_file_path)
        elif rule.get("name") == "BIOS ID info":
            analyze_bios_id_info(keyword, result, rule, worksheet, row_num, log_file_path)
        else:
            print(f"Rule not found in JSON file: {rule.get('name')}")

    # Set fixed column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 35
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 60

    # Adjust row heights based solely on the content length in column D
    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
        cell = row[0]
        content_length = len(str(cell.value)) if cell.value else 0
        worksheet.row_dimensions[cell.row].height = max(15, min(content_length // 10 * 15, 100))

    workbook.save(output_file)
    print(f"Analysis results saved to {output_file}")

# Main function to process the log file
def process_log_file(log_file_path):
    results = analyze_log_file(log_file_path, checks)
    save_to_excel(results, "output.xlsx", log_file_path)
    print(f"Analysis results saved to output.xlsx")

# Input log file path
log_file_path = 'SGX_enablement_putty.log'
process_log_file(log_file_path)
