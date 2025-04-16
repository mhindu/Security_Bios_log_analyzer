import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import filedialog

# Function to search for a keyword in a log file and extract the next bits
def search_keyword_in_log(log_file_path, keyword, num_bits, skip_bits=0, exclude_chars=None):
    next_bits = None
    with open(log_file_path, 'r') as file:
        lines = file.readlines()
    
    for line in lines:
        if keyword in line:
            # Extract the next bits after skipping the specified number of bits
            start_index = line.find(keyword) + len(keyword) + skip_bits
            next_bits = line[start_index:start_index + num_bits].strip() if num_bits > 0 else line[start_index:].strip()
            if exclude_chars:
                for char in exclude_chars:
                    next_bits = next_bits.replace(char, '')
    return next_bits
    

# Function to save DataFrame to an Excel file
def dataframe_to_excel(df, excel_file):
    df.to_excel(excel_file, index=False, engine='openpyxl')

# File paths
#log_file = 'input.log'
# Function to process the log file and save the results to an Excel file
def process_log_file(log_file_path):
    excel_file = 'output.xlsx'
    keywords = [
        ("[IA32_TME_ACTIVATE MSR 982h] =", 0),
        ("[IA32_TME_CAPABILITY MSR 981h] =", 0),
        ("MKTME: total keys", 3),
        ("mktme keys-ids", 3),
        ("ACTM location found!", 0, 0, ['(', ')']),  # (keyword, num_bits, skip_bits)
        ("starts ACTM launch (GETSEC[ENTERACCS])", 0),  # (keyword, num_bits)
        ("[ACTM] Socket[00] (SBSP) returned", 0, 0, ['(', ')']),  # (keyword, num_bits, skip_bits, exclude_chars)
        ("[LT EMIF register at 0xFED30200]: ", 10),
        ("Error: Mcheck code error -> LP0000", 0, 0, ['(', ')']),  # (keyword, num_bits, skip_bits, exclude_chars)
        ("SGX enabled", 0),  # (keyword, num_bits)
        ("build_date ", 8),  # (keyword, num_bits)
        ("build_num ", 3),  # (keyword, num_bits)
        ("TDX module initialized", 0),  # (keyword, num_bits)
        ("svtdx: Driver and seamrr target initialized", 0),  # (keyword, num_bits)
        ("LT_STATUS[0xFED30000]", 21) , # (keyword, num_bits)
        ("LT_EXTENDED_STATUS[0xFED30008]", 21) , # (keyword, num_bits)
        ("LT_BOOT_STATUS[0xFED300A0]", 21),  # (keyword, num_bits)
        ("LT_ERROR_CODE[0xFED30328]", 21) , # (keyword, num_bits)
        ("LT_CRASH[0xFED30030]", 21), # (keyword, num_bits)
        ("MSR_BOOT_GUARD_SACM_INFO[0x0000013A]", 21) , # (keyword, num_bits)
        ("Bios ID: ", 0)  # (keyword, num_bits)
    ]

    # Search for the keywords in the log file and get the next bits
    data = []
    for keyword, num_bits, *options in keywords:
        skip_bits = options[0] if len(options) > 0 else 0
        exclude_chars = options[1] if len(options) > 1 else None
        next_bits = search_keyword_in_log(log_file_path, keyword, num_bits, skip_bits, exclude_chars)
        data.append((keyword, next_bits))

    # Create a DataFrame with the keywords and next bits
    df = pd.DataFrame(data, columns=['Keyword', 'Next Bits'])

    # Save DataFrame to Excel
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1, header=False)
        worksheet = writer.sheets['Sheet1']
        for i, (keyword, next_bits) in enumerate(data, start=2):
            worksheet[f'B{i}'] = keyword if i != 10 else "ACTM ErrorCode"
            worksheet[f'C{i}'] = next_bits

    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    # Set the column width 
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 35
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 60


    # Add the specific cases for the given cells and set bold font
    bold_font = Font(bold=True)

    worksheet['A1'] = "Security Features"
    worksheet['A1'].font = bold_font
    worksheet['B1'] = "Feature check"
    worksheet['B1'].font = bold_font
    worksheet['C1'] = "Results from log"
    worksheet['C1'].font = bold_font
    worksheet['D1'] = "Analysis from Result"
    worksheet['D1'].font = bold_font
    worksheet['A2'] = "MKTME info"
    worksheet['A2'].font = bold_font
    worksheet['A6'] = "ACTM info"
    worksheet['A6'].font = bold_font
    worksheet['A10'] = "SGX info"
    worksheet['A10'].font = bold_font
    worksheet['A12'] = "TDX info"
    worksheet['A12'].font = bold_font
    worksheet['A18'] = "BTG & TXT info"
    worksheet['A18'].font = bold_font
    worksheet['A24'] = "System info"
    worksheet['A24'].font = bold_font


    # Add logic for IA32_TME_ACTIVATE MSR check
    def process_ia32_tme_activate(ia32_tme_activate_text, worksheet):
        ia32_tme_activate_text = ia32_tme_activate_text.strip().replace(" ", "")
        # Pad the hexadecimal value to 16 characters
        if ia32_tme_activate_text.startswith("0x"):
            ia32_tme_activate_text = "0x" + ia32_tme_activate_text[2:].zfill(16)
        else:
            ia32_tme_activate_text = ia32_tme_activate_text.zfill(16)
    
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(ia32_tme_activate_text):
            worksheet['C2'] = ia32_tme_activate_text
            ia32_tme_activate_value = bin(int(ia32_tme_activate_text, 16))[2:].zfill(64)
            result_text = []

            bits_32_to_35 = ia32_tme_activate_value[-36:-32]
            bits_32_to_35_hex = hex(int(bits_32_to_35, 2))
            bits_32_to_35_dec = 2 ** (int(bits_32_to_35, 2))
            result_text.append(f"Total key bits: {bits_32_to_35_hex} - ie.Total {bits_32_to_35_dec} keys are allocated")

            bit_1 = ia32_tme_activate_value[-2:-1]
            bits_1_hex = hex(int(bit_1, 2))
            result_text.append(f"TME bit: {bits_1_hex} - ie. key0 is allocated")
            
            bits_36_to_39 = ia32_tme_activate_value[-40:-36]
            bits_36_to_39_hex = hex(int(bits_36_to_39, 2))
            result_text.append(f"TDX key bits: {bits_36_to_39_hex} ")

            bits_48_to_63 = ia32_tme_activate_value[-64:-48]
            bits_48_to_63_hex = hex(int(bits_48_to_63, 2))
            alg_map = {
                '0x4': "AES-XTS-256 alg is enabled",
                '0xc': "AES-XTS-256i alg is enabled",
                '0x3': "AES-XTS-128i alg is enabled",
                '0x1': "AES-XTS-128 alg is enabled"
            }
            result_text.append(f"mktme_alg_enable: {bits_48_to_63_hex} ie. {alg_map.get(bits_48_to_63_hex, 'Not valid algorithm')}")

            result_text_str = "\n".join(result_text)
            worksheet['D2'].value = result_text_str
            worksheet['D2'].alignment = Alignment(wrapText=True)
        else:
            worksheet['C2'] = f"improper print - {ia32_tme_activate_text}"

    worksheet['B2'] = "TME_ACTIVATE MSR 982h"
    ia32_tme_activate_text = search_keyword_in_log(log_file_path, "[IA32_TME_ACTIVATE MSR 982h] =", 0)
    if ia32_tme_activate_text is None:
        ia32_tme_activate1_text = search_keyword_in_log(log_file_path, "MKTME: Reading IA32_TME_ACTIVATE MSR: ", 0)
        if ia32_tme_activate1_text is None:
            worksheet['C2'] = "TME_activate status is not present in log"
        else:
            if ia32_tme_activate1_text == "0x1":
                worksheet['C2'] = "Only TME bit is enabled"
                worksheet['D2'] = "Only TME bit is enabled"
            else:
                process_ia32_tme_activate(ia32_tme_activate1_text, worksheet)
    else:
        process_ia32_tme_activate(ia32_tme_activate_text, worksheet)


    # Add logic for IA32_TME_CAPABILITY MSR check

    def process_ia32_tme_capability(ia32_tme_capability_text, worksheet):
        ia32_tme_capability_text = ia32_tme_capability_text.strip().replace(" ", "")
        # Pad the hexadecimal value to 16 characters
        if ia32_tme_capability_text.startswith("0x"):
            ia32_tme_capability_text = "0x" + ia32_tme_capability_text[2:].zfill(16)
        else:
            ia32_tme_capability_text = ia32_tme_capability_text.zfill(16)

        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(ia32_tme_capability_text):
            worksheet['C3'] = ia32_tme_capability_text
            ia32_tme_capability_text = bin(int(ia32_tme_capability_text, 16))[2:].zfill(64)
            result_text = []

            bit_32_to_35 = ia32_tme_capability_text[-36:-32]
            bits_32_to_35_hex = hex(int(bit_32_to_35, 2))
            bits_32_to_35_int = 2 ** int(bit_32_to_35, 2)
            result_text.append(f"mktme_max_keyid_bits: {bits_32_to_35_hex} ie. {bits_32_to_35_int} keys")
            
            bits_36_to_50 = ia32_tme_capability_text[-51:-36]
            bits_36_to_50_hex = hex(int(bits_36_to_50, 2))
            bits_36_to_50_int = int(bits_36_to_50, 2)
            result_text.append(f"mktme_max_keys: {bits_36_to_50_hex} ie. [0-{bits_36_to_50_int} keys]")

            result_text_str = "\n".join(result_text)
            worksheet['D3'].value = result_text_str
            worksheet['D3'].alignment = Alignment(wrapText=True)
        else:
            worksheet['C3'] = f"improper print - {ia32_tme_capability_text}"

    worksheet['B3'] = "TME_CAPABILITY MSR 981h"
    ia32_tme_capability_text = search_keyword_in_log(log_file_path, "[IA32_TME_CAPABILITY MSR 981h] =", 0)
    if ia32_tme_capability_text is None:
        ia32_tme_capability_text = search_keyword_in_log(log_file_path, "MKTME: Reading IA32_TME_CAPABILITY MSR: ", 0)
        if ia32_tme_capability_text is None:
            worksheet['C3'] = "TME_Capability status is not present in log"
        else:
            process_ia32_tme_capability(ia32_tme_capability_text, worksheet)
    else:
        process_ia32_tme_capability(ia32_tme_capability_text, worksheet)

    # Add logic for MKTME: total keys check
    worksheet['B4'] = "Total Keys"
    total_keys_text = search_keyword_in_log(log_file_path, "MKTME: total keys", 3)
    if total_keys_text is None:
        worksheet['C4'] = "Keys allocation is not present in log"
    else:
        total_keys_text = total_keys_text.strip().replace(" ", "")
        result_text = [f"Total keys allocated: key0 - key{total_keys_text}"]
        worksheet['C4'] = total_keys_text
        worksheet['D4'].value = "\n".join(result_text)
        worksheet['D4'].alignment = Alignment(wrapText=True)

    # Add logic for mktme key-ids check
    worksheet['B5'] = "MKTME Keys"
    mktme_key_text = search_keyword_in_log(log_file_path, "mktme keys-ids", 3)
    if mktme_key_text is None:
        worksheet['C5'] = "MKTME keys are not enabled"
    else:
        mktme_key_text = mktme_key_text.strip().replace(" ", "")
        mktme_keys_int = int(mktme_key_text) + 1
        worksheet['C5'] = mktme_key_text
        result_text = []
        result_text.append(f"Mktme keys allocated: key1 - key{mktme_key_text}")
        #TDX key allocated
        if total_keys_text == mktme_key_text:
            result_text.append(f"No TDX keys are allocated")
        else:
            result_text.append(f"TDX keys allocated: key{mktme_keys_int} - key{total_keys_text}")
        worksheet['D5'].value = "\n".join(result_text)
        worksheet['D5'].alignment = Alignment(wrapText=True)
        
        
    # Add logic for ACTM location check
    worksheet['B6'] = "ACTM location"
    actm_location_text = search_keyword_in_log(log_file_path, "ACTM location found!", 12, 0, ['(', ')'])
    if actm_location_text is None:
        worksheet['C6'] = "ACTM location is not present in log"
    else:
        actm_location_text = actm_location_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]+$')
        if hex_pattern.match(actm_location_text):
            worksheet['C6'] = actm_location_text
        else:
            worksheet['C6'] = f"improper print - {actm_location_text}"

    # Add logic for start ACTM check
    worksheet['B7'] = "Is ACTM launch started?"
    actm_launch_text = search_keyword_in_log(log_file_path, "starts ACTM launch (GETSEC[ENTERACCS])", 0)
    if actm_launch_text is None:
        actm_launch1_text = search_keyword_in_log(log_file_path, "[ACTM] Launch skipped!", 0)
        if actm_launch1_text is None:
            worksheet['C7'] = "ACTM launch status is not present in log"
        else:
            worksheet['C7'] = "ACTM launch is skipped"
    else:
        worksheet['C7'] = "ACTM launch started"
        worksheet['D7'] = "ACTM launch started"

    # Add logic for ACTM error code check
    worksheet['B8'] = "ACTM ErrorCode"
    actm_error_code_text = search_keyword_in_log(log_file_path, "[ACTM] Socket[00] (SBSP) returned", 0, 0, ['(', ')'])
    if actm_error_code_text is None:
        worksheet['C8'] = "ACTM errorcode is not present in log"
    else:
        # Trim any leading or trailing whitespace
        actm_error_code_text = actm_error_code_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]+$')
        if hex_pattern.match(actm_error_code_text):
            worksheet['C8'] = actm_error_code_text
        else:
            worksheet['C8'] = f"improper print - {actm_error_code_text}"


    # Add logic for prod_sku check
    worksheet['B9'] = "Is this PROD SKU?"
    prod_sku_text = search_keyword_in_log(log_file_path, "[LT EMIF register at 0xFED30200]: ", 10)
    if prod_sku_text is None:
        worksheet['C9'] = "Prod SKU status is not present in log"
    else:
        # Trim any leading or trailing whitespace
        prod_sku_text = prod_sku_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]+$')
        if hex_pattern.match(prod_sku_text):
            worksheet['C9'] = prod_sku_text
            # Convert the hexadecimal value to binary
            prod_sku_binary = bin(int(prod_sku_text, 16))[2:].zfill(32)
            # Check if the 31st bit is set
            if prod_sku_binary[-32] == '1':
                worksheet['D9'] = "Bit 31 is set, It is prod SKU, recommended Prod patches"
            else:
                worksheet['D9'] = "Bit 31 is unset, It is debug SKU, recommended debug patches"
        else:
            worksheet['C9'] = f"improper print - {prod_sku_text}"
    
        

    # Add logic for SGX enabled check
    error_code_file_path = "Mcheck_error_codes.xlsx"
    worksheet['B10'] = "Mcheck Error Code ?"
    mcheck_code_text = search_keyword_in_log(log_file_path, "Error: Mcheck code error -> LP0000", 0 , 0, ['(', ')'])
    if mcheck_code_text is None:
        mcheck_code1_text = search_keyword_in_log(log_file_path, "[SGX] SgxErrorCode = ",0)
        if mcheck_code1_text == '0x0':
            worksheet['C10'] = mcheck_code1_text
            worksheet['D10'] = "No Mcheck - SGX enabled successfully"
        else:
            worksheet['C10'] = "mcheck error code is not present in log"
    else:
        # Trim any leading or trailing whitespace and remove 'F' just next to '0x'
        mcheck_code_text = mcheck_code_text.strip()
        if mcheck_code_text.startswith("0xF"):
            mcheck_code_text = "0x" + mcheck_code_text[3:]
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]+$')
        if hex_pattern.match(mcheck_code_text):
            worksheet['C10'] = mcheck_code_text
            # Search mcheck_code_text in the error_code.xlsx file
            error_wb = openpyxl.load_workbook(error_code_file_path)
            error_ws = error_wb.active

            found = False
            for row in error_ws.iter_rows(min_row=2, values_only=True):
                error_code = str(row[0]).strip().lower()  # Convert error code to lowercase
                if error_code == mcheck_code_text.lower():  # Convert mcheck_code_text to lowercase
                    worksheet['D10'] =f"Mcheck Error is " + row[1]
                    found = True
                    break

            if not found:
                worksheet['D10'] = "Error code not found in error_codes.xlsx"
        else:
            worksheet['C10'] = f"improper print - {mcheck_code_text}"

    # Add logic for SGX enabled check
    worksheet['B11'] = "Is SGX enabled?"
    sgx_enabled_text = search_keyword_in_log(log_file_path, "SGX (Secure Enclaves) feature enabled", 0)
    if sgx_enabled_text is None:
        worksheet['C11'] = "SGX is not enabled"
    else:
        worksheet['C11'] = "SGX is enabled"
        worksheet['D11'] = "SGX is enabled"
   

    # Add logic for TDX_build_date check
    worksheet['B12'] = "TDX module build date"
    tdx_build_date_text = search_keyword_in_log(log_file_path, "build_date ", 8)
    if tdx_build_date_text is None:
        worksheet['C12'] = "TDX is not enabled"
    else:
        worksheet['C12'] = tdx_build_date_text
        worksheet['D12'] = tdx_build_date_text

    # Add logic for TDX_build_num check
    worksheet['B13'] = "TDX module build version"
    tdx_build_num_text = search_keyword_in_log(log_file_path, "build_num ", 3)
    if tdx_build_num_text is None:
        worksheet['C13'] = "TDX is not enabled"
    else:
        worksheet['C13'] = tdx_build_num_text
        worksheet['D13'] = tdx_build_num_text

    worksheet['B14'] = "Is TDX module initialized?"
    tdx_initialized_text = search_keyword_in_log(log_file_path, "TDX module initialized", 0)
    if tdx_initialized_text is None:
        worksheet['C14'] = "TDX module not initialized"
    else:
        worksheet['C14'] = "TDX module initialized"
        worksheet['D14'] = "TDX module initialized"

    # Add logic for SEAMRR initialization check
    worksheet['B15'] = "Is SEAMRR Initialized?"
    seamrr_initialized_text = search_keyword_in_log(log_file_path, "svtdx: Driver and seamrr target initialized", 0)
    if seamrr_initialized_text is None:
        worksheet['C15'] = "SEAMRR target not found"
    else:
        worksheet['C15'] = "SEAMRR initialized"
        worksheet['D15'] = "SEAMRR initialized"


    # Add logic for SEAMRR base range check
    worksheet['B16'] = "SEAMRR_BASE MSR 1400h"
    seamrr_base_text = search_keyword_in_log(log_file_path, "SEAMRR_BASE MSR 1400h", 25)
    if seamrr_base_text is None:
        worksheet['C16'] = "SEAMRR Base range is not present in log"
    else:
        seamrr_base_text_cleaned = seamrr_base_text.replace("]", "").replace("=", "")
        # Trim any leading or trailing whitespace
        seamrr_base_text_cleaned = seamrr_base_text_cleaned.strip().replace(" ", "")
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        
        if hex_pattern.match(seamrr_base_text_cleaned):
            worksheet['C16'] = seamrr_base_text_cleaned
            # Convert the cleaned text to an integer
            seamrr_base_value = int(seamrr_base_text_cleaned, 16)
            # Check if bit 3 is set
            if seamrr_base_value & (1 << 3):
                worksheet['D16'] = f"Bit 3 is set, Bios configured SEAMRR_Base"
            else:
                worksheet['D16'] = f"Bit 3 is not set, Bios doesnot configured SEAMRR_Base"
        else:
            worksheet['C16'] = f"improper print - {seamrr_base_text_cleaned}"

    # Add logic for SEAMRR Mask range check
    worksheet['B17'] = "SEAMRR_MASK MSR 1401h"
    seamrr_mask_text = search_keyword_in_log(log_file_path, "SEAMRR_MASK MSR 1401h", 25)
    if seamrr_mask_text is None:
        worksheet['C17'] = "SEAMRR Mask range is not present in log"
    else:
        seamrr_mask_text_cleaned = seamrr_mask_text.replace("]", "").replace("=", "")
        # Trim any leading or trailing whitespace
        seamrr_mask_text_cleaned = seamrr_mask_text_cleaned.strip().replace(" ", "")
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(seamrr_mask_text_cleaned):
            worksheet['C17'] = seamrr_mask_text_cleaned
            bit10_set = seamrr_base_value & (1 << 10)
            bit11_set = seamrr_base_value & (1 << 11)
            
            if bit10_set and bit11_set:
                worksheet['D17'] = f"bit 10 lock bit and bit 11 valid bit are set"
            elif bit10_set:
                worksheet['D17'] = f"only bit 10 lock bit is set"
            elif bit11_set:
                worksheet['D17'] = f"only bit 11 valid is set"
            else:
                worksheet['D17'] = f"neither bit 10 nor bit 11 is set on SEAMRR_MASK"
        else:
            worksheet['C17'] = f"improper print - {seamrr_mask_text_cleaned}"


    # Add logic for lt_status info
    worksheet['B18'] = "LT_STATUS[0xFED30000]"
    lt_status_info_text = search_keyword_in_log(log_file_path, "LT_STATUS[0xFED30000]", 21, exclude_chars=['='])
    if lt_status_info_text is None:
        worksheet['C18'] = "LT_status is not present in log"
    else:
        # Trim any leading or trailing whitespace
        lt_status_info_text = lt_status_info_text.strip()
        # Regular expression for 16-digit hex value with optional 0x prefix
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(lt_status_info_text):
            worksheet['C18'] = lt_status_info_text
            lt_status_info_value = int(lt_status_info_text, 16)
            
            if lt_status_info_value & (1 << 1):
                worksheet['D18'] = f"bit 1 is set - SEXIT Status is Done"
            else:
                worksheet['D18'] = f"bit 1 is not set - SEXIT status is not Done"
        else:
            worksheet['C18'] = f"improper print - {lt_status_info_text}"

    # Add logic for lt_extended_status info
    worksheet['B19'] = "LT_EXTENDED_STATUS[0xFED30008]"
    lt_extended_status_info_text = search_keyword_in_log(log_file_path, "LT_EXTENDED_STATUS[0xFED30008]", 21, exclude_chars=['='])
    if lt_extended_status_info_text is None:
        worksheet['C19'] = "LT_extended_status is not present in log"
    else:
        # Trim any leading or trailing whitespace
        lt_extended_status_info_text = lt_extended_status_info_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(lt_extended_status_info_text):
            worksheet['C19'] = lt_extended_status_info_text
            lt_extended_status_info_value = int(lt_extended_status_info_text, 16)
            bit0_set = lt_extended_status_info_value & (1 << 0)
        
            if bit0_set:
                worksheet['D19'] = f"bit 0 is set - LT_poison_status is received"
            else:
                worksheet['D19'] = f"bit 0 is unset - No poison cycle is received"
        else:
            worksheet['C19'] = f"improper print - {lt_extended_status_info_text}"

    # Add logic for lt_boot_status info
    worksheet['B20'] = "LT_BOOT_STATUS[0xFED300A0]"
    lt_boot_status_info_text = search_keyword_in_log(log_file_path, "LT_BOOT_STATUS[0xFED300A0]", 21, exclude_chars=['='])
    if lt_boot_status_info_text is None:
        worksheet['C20'] = "LT_boot_status is not present in log"
    else:
        # Trim any leading or trailing whitespace
        lt_boot_status_info_text = lt_boot_status_info_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(lt_boot_status_info_text):
            worksheet['C20'] = lt_boot_status_info_text
            lt_boot_status_info_text = int(lt_boot_status_info_text, 16)
            bit30_set = lt_boot_status_info_text & (1 << 30)
            bit31_set = lt_boot_status_info_text & (1 << 31)
            bit59_set = lt_boot_status_info_text & (1 << 59)
            bit62_set = lt_boot_status_info_text & (1 << 62)
            bit63_set = lt_boot_status_info_text & (1 << 63)
            # Initialize the cell value
            result_text = []
            
            if bit30_set:
                result_text.append("Bit 30 set - TXT startup success")
            else:
                result_text.append("Bit 30 is unset - TXT startup not success")
            
            if bit31_set:
                result_text.append("Bit 31 set - BTG startup success")
            else:
                result_text.append("Bit 31 is unset - BTG startup not success")
            
            if bit59_set:
                result_text.append("Bit 59 set - BIOS is trusted as a part of TXT or BTG flow")
            else:
                result_text.append("Bit 59 is unset - BIOS is untrusted")
            
            if bit62_set:
                result_text.append("Bit 62 set - CPU sets error as SACM authentication failed")
            else:
                result_text.append("Bit 62 is unset - No CPU error - No SACM authentication failure")
            
            if bit63_set:
                result_text.append("Bit 63 set - S-ACM success as BTG/TXT flow is successfull ")
            else:
                result_text.append("Bit 63 is unset - SACM execution failed")
            
            # Join the list into a single string with newline characters
            result_text_str = "\n".join(result_text)
            worksheet['D20'].value = result_text_str
            worksheet['D20'].alignment = Alignment(wrapText=True)

        else:
            worksheet['C20'] = f"improper print - {lt_boot_status_info_text}"

    # Add logic for lt_error_code info
    worksheet['B21'] = "LT_ERROR_CODE[0xFED30328]"
    lt_error_code_info_text = search_keyword_in_log(log_file_path, "LT_ERROR_CODE[0xFED30328]", 21, exclude_chars=['='])
    if lt_error_code_info_text is None:
        worksheet['C21'] = "LT_error_status is not present in log"
    else:
        # Trim any leading or trailing whitespace
        lt_error_code_info_text = lt_error_code_info_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(lt_error_code_info_text):
            worksheet['C21'] = lt_error_code_info_text
            lt_error_code_info_value = int(lt_error_code_info_text, 16)
            # Initialize the cell value
            result_text = []
            
            if lt_error_code_info_value & (1 << 31):
                result_text.append("Bit 31 is set - Entire content of this register is valid")
                if lt_error_code_info_value & (1 << 15):
                    result_text.append("Bit 15 is set - SACM has started successfully")
                else:
                    result_text.append("Bit 15 is unset - SACM didnt start")

                if lt_error_code_info_value & (1 << 30):
                    result_text.append("Bit 30 is set - SACM has generated errorcodes")
                else:
                    result_text.append("Bit 30 is unset - No errorcode generated by SACM")
            else:
                result_text.append("Bit 31 is unset - Entire content of this register from bits 30:0 is invalid")
            
            # Join the list into a single string with newline characters
            result_text_str = "\n".join(result_text)
            worksheet['D21'].value = result_text_str
            worksheet['D21'].alignment = Alignment(wrapText=True)
        else:
            worksheet['C21'].value +=  f"improper print - {lt_error_code_info_text}"


    # Add logic for lt_crash info
    worksheet['B22'] = "LT_CRASH[0xFED30030] "
    lt_crash_info_text = search_keyword_in_log(log_file_path, "LT_CRASH[0xFED30030]", 21, exclude_chars=['='])
    if lt_crash_info_text is None:
        worksheet['C22'] = "LT_crash is not present in log"
    else:
        # Trim any leading or trailing whitespace
        lt_crash_info_text = lt_crash_info_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(lt_crash_info_text):
            worksheet['C22'] = lt_crash_info_text
            lt_crash_info_value = int(lt_crash_info_text, 16)
            
            if lt_crash_info_value & (1 << 31):
                worksheet['D22'] = f"Bit 31 is set - LT_Crash is set,"
                if lt_crash_info_value & (1 << 30):
                    worksheet['D22'].value += " Crash induced by internal CPU"
                else:
                    worksheet['D22'].value += " Crash induced by external ACM, MLE"
            else:
                worksheet['D22'] = f"Bit 31 is unset - No LT_Crash is seen"
        else:
            worksheet['C22'] = f"improper print - {lt_crash_info_text}"


    #add logic for sacm_info
    worksheet['B23'] = "SACM_INFO[0x0000013A]"
    sacm_info_text = search_keyword_in_log(log_file_path, "MSR_BOOT_GUARD_SACM_INFO[0x0000013A]", 21, exclude_chars=['='])
    if sacm_info_text is None:
        worksheet['C23'] = "SACM_INFO is not present in log"
    else:
        # Trim any leading or trailing whitespace
        sacm_info_text = sacm_info_text.strip()
        hex_pattern = re.compile(r'^(0x)?[0-9A-Fa-f]{16}$')
        if hex_pattern.match(sacm_info_text):
            worksheet['C23'] = sacm_info_text
            sacm_info_value = int(sacm_info_text, 16)
            bit3_set = sacm_info_value & (1 << 3)
            bit34_set = sacm_info_value & (1 << 34)
            bit32_set = sacm_info_value & (1 << 32)
            bit0_set = sacm_info_value & (1 << 0)
            bit2_set = sacm_info_value & (1 << 2)
            bit4_set = sacm_info_value & (1 << 4)
            bit5_set = sacm_info_value & (1 << 5)
            bit6_set = sacm_info_value & (1 << 6)
            # Initialize the cell value
            result_text = []
            
            if bit34_set:
                result_text.append("Bit 34 set - LT_SX_EN Fuse is enabled")
            else:
                result_text.append("Bit 34 is unset - LT_SX_EN Fuse is disabled")
            
            if bit32_set:
                result_text.append("Bit 32 set - BTG is enabled")

                if bit0_set and bit2_set and bit3_set and bit4_set and bit5_set and bit6_set:
                    result_text.append("BTG-5 profile is enabled & TPM Success")

                elif bit0_set and bit2_set and bit3_set and bit5_set and bit6_set:
                    result_text.append("BTG-3 profile is enabled & TPM Success")

                elif bit0_set and bit4_set and bit6_set:
                    result_text.append("BTG-4 profile is enabled")

                else:
                    result_text.append("No BTG profile enabled")
            else:
                result_text.append("Bit 32 is unset - BTG is not enabled")
            
            # Join the list into a single string with newline characters
            result_text_str = "\n".join(result_text)
            worksheet['D23'].value = result_text_str
            worksheet['D23'].alignment = Alignment(wrapText=True)
        else:
            worksheet['C23'] = f"improper print - {sacm_info_text}"

    # Add logic for system info
    worksheet['B24'] = "BIOS ID"
    Bios_info_text = search_keyword_in_log(log_file_path, "Bios ID: ", 0)
    if Bios_info_text is None:
        worksheet['C24'] = "Bios_ID is not present in log"
    else:
        worksheet['C24'] = Bios_info_text

    # Clear the specified cells
    for cell in ['A3', 'A4', 'A5', 'A7', 'A8', 'A9', 'A11', 'A13', 'A14', 'A15', 'A16','A17','A19','A20','A21','A22','A23']:
        worksheet[cell] = ""


    # Save the workbook
    workbook.save(excel_file)

    print(f"Data from {log_file_path} has been successfully written to {excel_file}.")


# Function to open the file dialog and select the log file
def open_file():
    global log_file_path
    log_file_path = filedialog.askopenfilename(filetypes=[("Log files", "*.log"), ("Text files", "*.txt")])
    if log_file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, log_file_path)
        status_label.config(text=f"Input file: '{log_file_path}'")

# Function to process the log file
def process_file():
    if log_file_path:
        process_log_file(log_file_path)

# Function to save the output Excel file to a user-specified location
def save_output():
    if log_file_path:
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            with open('output.xlsx', 'rb') as fsrc:
                with open(save_path, 'wb') as fdst:
                    fdst.write(fsrc.read())
            status_label.config(text=f"Output saved as {save_path}")
            save_entry.delete(0, tk.END)
            save_entry.insert(0, save_path)

# Create the GUI
root = tk.Tk()
root.title("Security Data Extractor from LOG")
root.geometry("600x400")  # Set the window size to 500x400 pixels
root.configure(bg="#f7f1f7")  # Set background color

# Set font styles
title_font = ("Helvetica", 20, "bold")
button_font = ("Helvetica", 12, "bold")
label_font = ("Helvetica", 12)

# Create and place the title label with a light background color
title_label = tk.Label(root, text="Security Data Analyzer from Bootlog", font=title_font, fg="white", bg="#660000")
title_label.pack(pady=20, padx=10, fill=tk.X)

# Create and place the buttons and label
browse_button = tk.Button(root, text="Browse Log File", command=open_file, font=button_font, bg="#028292", fg="white")
browse_button.pack(pady=10)

# Create and place the file entry box
file_entry = tk.Entry(root, font=label_font, width=50)
file_entry.pack(pady=10)

process_button = tk.Button(root, text="Process Security Data", command=process_file, font=button_font, bg="#C98391", fg="white")
process_button.pack(pady=10)

save_button = tk.Button(root, text="Save Output XL", command=save_output, font=button_font, bg="#FFC125", fg="white")
save_button.pack(pady=10)

# Create and place the save entry box
save_entry = tk.Entry(root, font=label_font, width=50)
save_entry.pack(pady=10)

exit_button = tk.Button(root, text="Exit", command=root.quit, font=button_font, bg="#f77268", fg="white")
exit_button.pack(pady=10)

status_label = tk.Label(root, text="", font=label_font, bg="#f7f7f7")
status_label.pack(pady=20)

root.mainloop()